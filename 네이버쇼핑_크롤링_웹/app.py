from flask import Flask, render_template, request
from openpyxl.styles.differential import DifferentialStyle
import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup
from selenium import webdriver

wb = Workbook()
write_ws = wb.active
app = Flask(__name__)


@app.route('/')
def hello():
    return render_template("index.html")

@app.route('/result', methods=['POST'])
def result():
    keyword = request.form['input1']
    page = int(request.form['input2'])
    titles = []
    links = []

    for i in range(1, page+1):
        driver = webdriver.Chrome("C:/Users/seob6/Desktop/chromedriver/chromedriver/chromedriver.exe")
        driver.implicitly_wait(3)
        driver.get(f"https://search.shopping.naver.com/search/all?frm=NVSHATC&origQuery={keyword}&pagingIndex={i}&pagingSize=20&productSet=total&query={keyword}&sort=rel&timestamp=&viewType=list")
        # 스크롤 내려가서 크롤링
        last_height = driver.execute_script("return document.body.scrollHeight")
        while 1:
            # scroll down to bottom
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            driver.implicitly_wait(2)

            # calculate new scroll height and compare with last scroll height
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                html = driver.page_source
                soup = BeautifulSoup(html, 'html.parser')
                title_links = soup.select("div.basicList_info_area__17Xyo > div.basicList_title__3P9Q7 > a")
                for title_link in title_links:
                    titles.append(title_link.text)
                    links.append(title_link['href'])
                break
            last_height = new_height
    # 엑셀 파일 저장
    for i, j in zip(range(1, len(titles) + 1), range(1, len(links) + 1)):
        write_ws.cell(i, 1, titles[i-1])
        write_ws.cell(j, 2, links[j-1])

    wb.save('static/crawl_result.xlsx')
    
    return render_template("result.html", titles=titles, links=links, zip=zip)


if __name__ == '__main__':
    app.run()

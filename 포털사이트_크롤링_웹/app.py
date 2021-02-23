from flask import Flask, render_template, request
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb = Workbook()
write_ws = wb.active
app = Flask(__name__)

@app.route('/')
def hello_world():
    return render_template("index.html")

@app.route('/result', methods=['POST'])
def result():
    keyword = request.form['input1']
    page = int(request.form['input2'])
    daum_list = []
    daum_url = []

    for i in range(1, page+1):
        req = requests.get(
            f"https://search.daum.net/search?nil_suggest=sugsch&w=news&DA=PGD&cluster=y&q={keyword}&p={i}")

        soup = BeautifulSoup(req.text, 'lxml')
        titles = soup.select("a.f_link_b")
        links = soup.select(".wrap_tit > a")

        for i in titles:
            print(i.text)
            daum_list.append(i.text)

        for link in links:
            daum_url.append(link["href"])

    # 엑셀 파일 저장
    for i in range(1, len(daum_list)+1):
        write_ws.cell(i, 1, daum_list[i-1])
    
    for j in range(1, len(daum_url)+1):
        write_ws.cell(j, 2, daum_url[j-1])

    wb.save('crawl_result.xlsx')

    return render_template("result.html", daum_list=daum_list, daum_url=daum_url, zip=zip)

if __name__ == '__main__':
    app.run()
